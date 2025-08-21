import openpyxl

def read_xlsx_as_dict_list(filepath,first_row=1, sheet_name=None):
    """
    Reads an XLSX file and returns its data as a list of dictionaries,
    similar to csv.DictReader output.

    Args:
        filepath (str): The path to the XLSX file.
        sheet_name (str, optional): The name of the sheet to read.
                                    If None, the active sheet is used.

    Returns:
        list: A list of dictionaries, where each dictionary represents a row
              and keys are column headers.
    """
    try:
        workbook = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        print(f"Error: File '{filepath}' not found.")
        return []

    if sheet_name:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            print(f"Error: Sheet '{sheet_name}' not found in '{filepath}'.")
            return []
    else:
        sheet = workbook.active

    data = []
    # Get column headers from the first row
    headers = [cell.value for cell in sheet[first_row]]

    # Iterate through rows starting from the second row (data rows)
    for row_index in range(first_row+1, sheet.max_row + 1):
        row_values = [cell.value for cell in sheet[row_index]]
        row_dict = dict(zip(headers, row_values))
        data.append(row_dict)

    return data



import csv
def ReadCSV(fname):
	with open(fname, newline='',encoding='utf-8-sig') as csvfile:
		data = csv.DictReader(filter(lambda row: row[0]!='#', csvfile))
		# data = csv.DictReader(csvfile)
		listdata=list(data)
		# for row in data:
		# 	print(row)
#        print(', '.join(row))
#    spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
	return listdata


# creo article.size a partir de articulo.descripcion.
# si tiene las palabras contenido o peso, lo borro o lo remplazo por lo que haya entre parentesis
#
import html
import re 
import math
def set_article(article):	
	#print(type(article["precio"]))
	print(article["title"])
	article["imagen"]=article["imagen"].lower()
	#Aumento el precio!
	aumento=int(article["precio"])*1 #.15
	redondeo=(math.ceil(aumento/100))*100
	# print(int(article["price"]),aumento,redondeo)
	article["precio"]=str(redondeo)
	#print(article["peso"], "gr")
	article["peso"]=int(article["peso"])
	# article['description']=article['descripcion']
	msj='Todavía no contamos con una descripción detallada del artículo'
	if 'description' not in article:
		article['description']=msj;
	else:
		if article['description']=='':
			article['description']=msj;

	article["descripcion"]=html.unescape(article["descripcion"])

	longString=article['descripcion']
	removal_list=['Contenido','Peso']

	if article['descripcion']!='':
		article['description']=article['descripcion']

	# if removal_list in longString:
	if next((True for word in removal_list if word in longString),False):		
		if '(' and ')' in longString:
			# https://stackoverflow.com/questions/38999344/extract-string-within-parentheses-python
			text_inside_paranthesis = re.findall('\(([^)]+)',longString)[0]
			removal_list.append('('+text_inside_paranthesis+')')
			#text_inside_paranthesis = re.findall('\(.*?\)',longString)[0]
			#removal_list.append(text_inside_paranthesis)
			article['descripcion']=text_inside_paranthesis;
		else:
			article['descripcion']=""
		# edit_string_as_list = longString.split()
		# print(removal_list)
		# final_list = [word for word in edit_string_as_list if word not in removal_list]
		# article["size"] = ' '.join(final_list)
		# print(final_list)
		for string in removal_list:
			# print(string)
			longString=longString.replace(string,'')
		longString=longString.replace('kilogramos','kg')
		longString=longString.replace('kilogramo','kg')
		longString=longString.replace('gramos','gr')
		# edit_string_as_list = longString.split('. ')
		edit_string_as_list = re.split('\. |\n',longString)
		longString=edit_string_as_list[0];
		# print(longString)

		if "size" not in article:
			article["size"] = longString.replace(':',"")
		if(len(edit_string_as_list)>1):
			edit_string_as_list.pop(0);
			for string in edit_string_as_list:
				if string =='':
					edit_string_as_list.remove(string)
			# if article['title']=='Grasa Bovina':
			# 	print(edit_string_as_list)	
			article['descripcion']=". ".join(edit_string_as_list);
	else:
		if 'Teléfono' in article:
			article["size"]	= article['Teléfono']
		else:
			if 'size' in article:
				if article['size'].isdigit():
					article['size']=str(article['size'])+' semillas'
			else:
				article["size"]=""
	#guardo el titulo en el id que es único
	article['id']=article['title'];

	# categories_list=article['categories'].split('>').Upper();
	categories_list=article['categorias'].split('>');
	
	article['title']=article['title'].lower()

	
	#al titulo le saco la categoria que tenga
	print('myfuncs.py -> categories_list')
	for w in categories_list:
		print(w)
		categoria=w.lower();
		if categoria in article['title'].lower():
			article['title']=article['title'].replace(categoria,'')	
		if categoria[-1]=='s':
			categoria=categoria[:-1]	
			if categoria in article['title'].lower():
				article['title']=article['title'].replace(categoria,'')		
	print(article["title"])
	print('-----')
	
	#al título le saco los paréntesis que tenga
	if '(' and ')' in article["title"]:
#		print(article["title"])
		# text_inside_paranthesis = re.findall('\(.*?\)',article["title"])[0]
		text_inside_paranthesis ="" # re.findall('\(([^)]+)',article["title"])[0]
		text_inside_paranthesis = re.findall('\(([^)]+)',article["title"])[0]
		article["title"]=(article["title"].replace('('+text_inside_paranthesis+')','')).rstrip()
		if(text_inside_paranthesis.lower() not in article['descripcion'].lower()):
			article["descripcion"]='<p>'+text_inside_paranthesis.upper()+ '</p> ' + article["descripcion"]


	Enpadronar(article)

	return article

def Enpadronar(article):
	exception_list=["nada te turbe","colores y aromas", "virgencita de luján", "sagrada familia", "virgen maría", "cola de rata", "24 glorias"]
	for w in exception_list:
		if article['parent'].lower() in w or article['title'].lower() in w or w in article['title'].lower():
			article['Nombre']=article['title'].upper()
			article['Apellido']=''
			return
	# if 'Nombre' in article:
	# 	if 'Apellido'  in article:
	# 		return
	# 	else:
	# 		article['Apellido']=''
	# 		return
	# if 'Apellido' in article:
	# 	article['Nombre']=''
	# 	return
	# if article['title']=='':
	# 	return

	if 'Nombre' in article:
		print("###########")
		print(article["Nombre"])
		if article['Nombre']!= '' and article['Nombre']!= None:
			return

	article['Nombre']=''
	lista=article['title'].split()
	print('2222222')
	if len(lista)==0:
		lista=['-1']
	print(lista)
	def set_name(article,lista):
	# la primer palabra siempre esta dentro del nombre
		if len(lista)>1:
			article['Nombre']+=' '+lista[0].upper()
		if len(lista)==1:
			article['Nombre']=lista[0].upper()
			article['Apellido']=''
			return 1
		if lista:
			lista.pop(0);
		return 0


	if set_name(article,lista):
		return

	print(lista)
	preposiciones=['de','con','en']
	if lista[0] in preposiciones:
# si la 2da palabra es 'de', 'con', ... entonces empieza el apellido
		article['Apellido']=' '.join(lista)
		return
	if len(lista)>1 and lista[1] in preposiciones:
		if set_name(article,lista):
			return
	while len(lista)>0 and lista[0].isupper():
# si la 2da palabra está en mayúscula entonces también pertenece al nombre
		article['Nombre'] += ' ' + lista[0]
		lista.pop(0);

	article['Apellido']=' '.join(lista)
	return

# import sys
# def join(f1,f2,fout):
def join(f2,f1=''):
	# if len(sys.argv) > 2:
	if f1:
		with open(f1) as fp:
			ff1 = fp.read()
	else :
		ff1= ''
	with open(f2) as fp:
		ff2 = fp.read()

	# with open (fout, 'w') as fp:
	with open (f2, 'w') as fp:
		fp.write(ff1+"\n"+ff2)

	return ff2
